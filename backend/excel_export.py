"""
Excel export for BayOps AI.

Layout uses fixed row anchors so the live-update function always knows
exactly which cells to refresh without re-reading the file:

  Row 1      — Title banner  (static, never changes)
  Row 2      — blank
  Rows 3-6   — Metadata labels (col A) + values (col B, updated live)
  Row 7      — blank
  Row 8      — "PARTS ORDERED" section header  (static)
  Row 9      — Parts column headers             (static)
  Row 10+    — Variable: parts rows, blank, labor section, totals
"""

import os
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe_print(msg: str):
    try:
        print(msg, flush=True)
    except (ValueError, OSError):
        pass


# Variable data always starts here; rows 1-9 are static structure
PARTS_START = 10

# openpyxl fill constants
_BLUE  = PatternFill("solid", fgColor="1E40AF")
_GRAY  = PatternFill("solid", fgColor="F1F5F9")
_GREEN = PatternFill("solid", fgColor="D1FAE5")
_ALT   = PatternFill("solid", fgColor="EFF6FF")
_WHITE = PatternFill("solid", fgColor="FFFFFF")
_THEAD = PatternFill("solid", fgColor="E2E8F0")


def _col(n: int) -> str:
    """1-indexed column → letter (1→A … 7→G)."""
    return chr(64 + n)


def _vehicle_str(bay_data: dict) -> str:
    veh = bay_data.get("vehicle") or {}
    return " ".join(filter(None, [veh.get("year"), veh.get("make"), veh.get("model")])) or "—"


# ── openpyxl helpers ───────────────────────────────────────────────────────────

def _write_static_header(ws):
    """Write rows 1-9 once at file creation. These never change."""
    for i, w in enumerate([32, 22, 14, 12, 12, 10, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 — blue title banner
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="BayOps AI — Service Order")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = _BLUE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Rows 3-6 — metadata labels (values written by _write_variable_data)
    for ri, label in enumerate(["Date:", "Bay Number:", "Technician:", "Vehicle:"], start=3):
        c = ws.cell(row=ri, column=1, value=label)
        c.font = Font(bold=True, color="475569")
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Row 8 — PARTS ORDERED section header
    ws.merge_cells("A8:G8")
    c = ws.cell(row=8, column=1, value="PARTS ORDERED")
    c.font = Font(bold=True, size=11)
    c.fill = _GRAY
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[8].height = 20

    # Row 9 — parts column headers
    for ci, h in enumerate(["Description", "Vendor", "Part #", "Qty", "Unit Cost", "Markup %", "Extended Price"], 1):
        c = ws.cell(row=9, column=ci, value=h)
        c.font = Font(bold=True, size=9, color="64748B")
        c.fill = _THEAD
        c.alignment = Alignment(horizontal="center" if ci > 3 else "left", vertical="center")


def _write_variable_data(ws, bay_data: dict):
    """Write metadata values (col B, rows 3-6) and all variable rows from PARTS_START."""
    veh_str = _vehicle_str(bay_data)
    for ri, val in enumerate([
        bay_data.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        bay_data.get("bay_number", "—"),
        bay_data.get("technician_name", "—"),
        veh_str,
    ], start=3):
        c = ws.cell(row=ri, column=2, value=str(val))
        c.font = Font(color="0F172A")
        c.alignment = Alignment(horizontal="left", vertical="center")

    row = PARTS_START

    # Parts rows
    for pi, part in enumerate(bay_data.get("parts", [])):
        fill = _ALT if pi % 2 == 0 else _WHITE
        vals = [
            part.get("description", ""),
            part.get("vendor", ""),
            part.get("part_number", "N/A"),
            part.get("quantity", 1),
            part.get("unit_cost", 0.0),
            f"{float(part.get('markup_pct', 25)):.0f}%",
            part.get("extended_price", 0.0),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.alignment = Alignment(
                horizontal="center" if ci in (4, 6) else ("right" if ci in (5, 7) else "left"),
                vertical="center",
            )
            if ci in (5, 7):
                c.number_format = "$#,##0.00"
        row += 1

    row += 1  # blank gap

    # Labor section
    labor = bay_data.get("labor", [])
    if labor:
        try:
            ws.merge_cells(f"A{row}:G{row}")
        except Exception:
            pass
        c = ws.cell(row=row, column=1, value="LABOR")
        c.font = Font(bold=True, size=11)
        c.fill = _GRAY
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        for ci, h in enumerate(["Description", "Hours", "Rate / Hr", "Total"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, size=9, color="64748B")
            c.fill = _THEAD
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        row += 1

        for li, item in enumerate(labor):
            fill = _ALT if li % 2 == 0 else _WHITE
            vals = [
                item.get("description", ""),
                item.get("hours", item.get("quantity", 0)),
                item.get("rate", item.get("unit_cost", 150)),
                item.get("extended_price", 0.0),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = fill
                c.alignment = Alignment(
                    horizontal="center" if ci == 2 else ("right" if ci in (3, 4) else "left"),
                    vertical="center",
                )
                if ci in (3, 4):
                    c.number_format = "$#,##0.00"
            row += 1

    row += 1  # blank before totals

    # Totals
    tax_pct = bay_data.get("tax_rate", 0.0825) * 100
    for label, val, is_grand in [
        ("Parts Subtotal",        bay_data.get("parts_subtotal", 0.0),  False),
        ("Labor Subtotal",        bay_data.get("labor_subtotal", 0.0),  False),
        (f"Tax ({tax_pct:.2f}%)", bay_data.get("tax_amount", 0.0),      False),
        ("GRAND TOTAL",           bay_data.get("grand_total", 0.0),     True),
    ]:
        try:
            ws.merge_cells(f"A{row}:F{row}")
        except Exception:
            pass
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=is_grand, color="0F172A" if is_grand else "475569")
        lc.alignment = Alignment(horizontal="right", vertical="center")
        if is_grand:
            lc.fill = _GREEN
            ws.row_dimensions[row].height = 22

        vc = ws.cell(row=row, column=7, value=val)
        vc.number_format = "$#,##0.00"
        vc.font = Font(bold=is_grand, color="059669" if is_grand else "0F172A")
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if is_grand:
            vc.fill = _GREEN
        row += 1


# ── Public API ─────────────────────────────────────────────────────────────────

def export_order_to_excel(bay_data: dict) -> str:
    """Create a new formatted .xlsx on the Desktop. Returns the absolute file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Service Order"
    _write_static_header(ws)
    _write_variable_data(ws, bay_data)

    desktop = Path.home() / "Desktop"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = str(desktop / f"BayOps_Order_{ts}.xlsx")
    wb.save(file_path)
    return file_path


def open_in_excel(file_path: str) -> bool:
    """Open the file in Excel via xlwings so we can drive it live later.
    Falls back to os.startfile if xlwings / Excel is unavailable."""
    try:
        import xlwings as xw
        xw.Book(file_path)
        return True
    except Exception:
        import subprocess
        subprocess.Popen(["cmd", "/c", "start", "", file_path], shell=False)
        return False


def update_excel_live(file_path: str, bay_data: dict):
    """
    Connect to the already-open Excel workbook and update all variable cells
    live — with visual cursor movement before each row so the user can watch
    the data change in real time (like a browser agent clicking around).

    Falls back to openpyxl file-rewrite if xlwings / Excel is unavailable.
    """
    filename = os.path.basename(file_path)

    try:
        import xlwings as xw

        # Locate the open workbook among all running Excel instances
        wb = None
        for xl_app in xw.apps:
            for book in xl_app.books:
                if book.name.lower() == filename.lower():
                    wb = book
                    break
            if wb:
                break

        if wb is None:
            # Workbook was closed — reopen it
            wb = xw.Book(file_path)

        ws = wb.sheets["Service Order"]
        wb.app.screen_updating = True   # ensure screen repaints in real time

        veh_str = _vehicle_str(bay_data)
        ts = bay_data.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # ── 1. Metadata cells (visual cursor jumps to each before writing) ────
        for addr, val in [
            ("B3", ts),
            ("B4", bay_data.get("bay_number", "—")),
            ("B5", bay_data.get("technician_name", "—")),
            ("B6", veh_str),
        ]:
            try:
                rng = ws.range(addr)
                rng.select()
                time.sleep(0.08)
                rng.value = str(val)
            except Exception:
                pass

        # ── 2. Clear all variable rows (values only — keeps static formatting) ─
        try:
            ws.range(f"A{PARTS_START}:G300").clear_contents()
        except Exception:
            pass

        row = PARTS_START

        # ── 3. Parts rows ─────────────────────────────────────────────────────
        # Use explicit A:G range so xlwings unambiguously writes one horizontal row
        for part in bay_data.get("parts", []):
            vals = [
                part.get("description", ""),
                part.get("vendor", ""),
                part.get("part_number", "N/A"),
                part.get("quantity", 1),
                part.get("unit_cost", 0.0),
                f"{float(part.get('markup_pct', 25)):.0f}%",
                part.get("extended_price", 0.0),
            ]
            try:
                ws.range(f"A{row}:G{row}").select()
                time.sleep(0.12)
                ws.range(f"A{row}:G{row}").value = vals   # explicit 1×7 row
                ws.range(f"E{row}").number_format = "$#,##0.00"
                ws.range(f"G{row}").number_format = "$#,##0.00"
            except Exception:
                pass
            row += 1

        row += 1  # blank gap

        # ── 4. Labor rows ──────────────────────────────────────────────────────
        labor = bay_data.get("labor", [])
        if labor:
            try:
                ws.range(f"A{row}").select()
                time.sleep(0.1)
                ws.range(f"A{row}").value = "LABOR"
            except Exception:
                pass
            row += 1

            # Re-write sub-headers (clear_contents above wiped them)
            try:
                ws.range(f"A{row}:D{row}").select()
                time.sleep(0.08)
                ws.range(f"A{row}:D{row}").value = ["Description", "Hours", "Rate / Hr", "Total"]
            except Exception:
                pass
            row += 1

            for item in labor:
                vals = [
                    item.get("description", ""),
                    item.get("hours", item.get("quantity", 0)),
                    item.get("rate", item.get("unit_cost", 150)),
                    item.get("extended_price", 0.0),
                ]
                try:
                    ws.range(f"A{row}:D{row}").select()
                    time.sleep(0.12)
                    ws.range(f"A{row}:D{row}").value = vals
                    ws.range(f"C{row}").number_format = "$#,##0.00"
                    ws.range(f"D{row}").number_format = "$#,##0.00"
                except Exception:
                    pass
                row += 1

        row += 1  # blank before totals

        # ── 5. Totals ──────────────────────────────────────────────────────────
        # Write label to A (home cell of merged A:F) and value to G separately.
        # Writing a list across a merged range in xlwings can silently unmerge
        # cells, causing label misalignment.
        tax_pct = bay_data.get("tax_rate", 0.0825) * 100
        for label, val in [
            ("Parts Subtotal",        bay_data.get("parts_subtotal", 0.0)),
            ("Labor Subtotal",        bay_data.get("labor_subtotal", 0.0)),
            (f"Tax ({tax_pct:.2f}%)", bay_data.get("tax_amount", 0.0)),
            ("GRAND TOTAL",           bay_data.get("grand_total", 0.0)),
        ]:
            try:
                ws.range(f"A{row}").select()
                time.sleep(0.12)
                ws.range(f"A{row}").value = label
                ws.range(f"G{row}").value = val
                ws.range(f"G{row}").number_format = "$#,##0.00"
            except Exception:
                pass
            row += 1

        wb.save()
        _safe_print(f"[Excel live] Updated {filename} — total ${bay_data.get('grand_total', 0):.2f}")

    except ImportError:
        _fallback_rewrite(file_path, bay_data)
    except Exception as e:
        _safe_print(f"[Excel live update error] {e}")
        _fallback_rewrite(file_path, bay_data)


def _fallback_rewrite(file_path: str, bay_data: dict):
    """Rewrite with openpyxl when xlwings / Excel COM is unavailable."""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Service Order"]
        for r in range(PARTS_START, ws.max_row + 1):
            for c in range(1, 8):
                ws.cell(row=r, column=c).value = None
        _write_variable_data(ws, bay_data)
        wb.save(file_path)
        _safe_print(f"[Excel fallback] Rewrote {os.path.basename(file_path)}")
    except Exception as ex:
        _safe_print(f"[Excel fallback error] {ex}")
